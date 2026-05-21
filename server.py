import sys, os, json, shutil, threading, webbrowser, sqlite3, uuid
from datetime import datetime
from flask import Flask, request, send_file, send_from_directory, jsonify
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

BASE = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
app = Flask(__name__, static_folder=os.path.join(BASE, 'static'))

# ===== DB =====
DB_PATH = os.path.join(os.path.expanduser('~'), '.usx_app', 'records.db')
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS records (
        id TEXT PRIMARY KEY, name TEXT NOT NULL,
        created_at TEXT NOT NULL, data TEXT NOT NULL)''')
    conn.commit(); conn.close()

init_db()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ===== Excel生成 =====
# info=ラベル列(G,K,O,S...) UC製造番号/冷却加熱/製造年/記録時間を書く
# no_val=号機番号列(H,L,P,T...)
# A/B/C/D=回路列
UNIT_COLS = [
    {"info":"G",  "no_val":"H",  "A":"G",  "B":"H",  "C":"I",  "D":"J"},
    {"info":"K",  "no_val":"L",  "A":"K",  "B":"L",  "C":"M",  "D":"N"},
    {"info":"O",  "no_val":"P",  "A":"O",  "B":"P",  "C":"Q",  "D":"R"},
    {"info":"S",  "no_val":"T",  "A":"S",  "B":"T",  "C":"U",  "D":"V"},
    {"info":"W",  "no_val":"X",  "A":"W",  "B":"X",  "C":"Y",  "D":"Z"},
    {"info":"AA", "no_val":"AB", "A":"AA", "B":"AB", "C":"AC", "D":"AD"},
    {"info":"AE", "no_val":"AF", "A":"AE", "B":"AF", "C":"AG", "D":"AH"},
    {"info":"AI", "no_val":"AJ", "A":"AI", "B":"AJ", "C":"AK", "D":"AL"},
    {"info":"AM", "no_val":"AN", "A":"AM", "B":"AN", "C":"AO", "D":"AP"},
    {"info":"AQ", "no_val":"AR", "A":"AQ", "B":"AR", "C":"AS", "D":"AT"},
    {"info":"AU", "no_val":"AV", "A":"AU", "B":"AV", "C":"AW", "D":"AX"},
    {"info":"AY", "no_val":"AZ", "A":"AY", "B":"AZ", "C":"BA", "D":"BB"},
]

FIELD_ROWS = [
    ("運転時間",            16, True),
    ("運転回数",            17, True),
    ("冷温水入口圧力",      18, False),
    ("冷温水出口圧力",      19, False),
    ("換算流量",            20, False),
    ("圧縮機電流_R",        21, True),
    ("圧縮機電流_S",        22, True),
    ("圧縮機電流_T",        23, True),
    ("圧縮機電圧_RS",       24, False),
    ("圧縮機電圧_ST",       25, False),
    ("圧縮機電圧_TR",       26, False),
    ("冷温水入口温度",      27, False),
    ("冷温水中間温度",      28, False),
    ("冷温水出口温度",      29, False),
    ("外気温度",            30, False),
    ("冷媒高圧圧力",        31, True),
    ("冷媒低圧圧力",        32, True),
    ("冷媒吐出ガス温度",    33, True),
    ("冷媒吸入ガス温度",    34, True),
    ("冷媒コイルガス温度1", 35, True),
    ("冷媒コイルガス温度2", 36, True),
    ("ファン回転数",        37, True),
    ("圧縮機運転周波数",    38, True),
    ("膨脹弁1開度",         39, True),
    ("膨脹弁2開度",         40, True),
    ("高圧圧力異常",        42, True),
    ("絶縁抵抗",            44, True),
    ("出荷時充填量",        46, True),
    ("初期充填量",          47, True),
    ("総充填量",            48, True),
]

CHECK_CELLS = {
    '圧縮機関係':       {'異常音':'F50','異常振動':'H50','圧力不良':'K50','オイル量':'N50','異常過熱':'Q50'},
    '凝縮器関係':       {'汚れ':'F51','流量不足':'H51','Yスト詰り':'K51','風量不足':'N51'},
    '蒸発器関係':       {'汚れ':'F52','流量不足':'H52','Yスト詰り':'K52','風量不足':'N52'},
    '送風機関係':       {'異常音':'F53','異常振動':'H53','ショートサイクル':'K53'},
    '冷媒配管系統部品関係': {'ストレーナ':'F54','四方弁':'H54','電磁弁':'K54','膨張弁':'N54'},
    '電装品関係':       {'センサー':'F55','リレー':'H55','トランス':'K55','制御基板':'N55','通信不具合':'Q55'},
    '補機関係':         {'ポンプ':'F56','圧力計':'H56','アキュムレータ':'K56'},
    'その他':           {'その他':'F57'},
}

PRECHECK_MAP = {
    '電装品結線状態確認':    ('F59','H59'),
    '通信線アドレス設定':    ('F60','H60'),
    '熱源機外観点検':        ('F61','H61'),
    '内蔵ポンプ動作確認':    ('F62','H62'),
    '水配管フラッシング':    ('Q59','T59'),
    '12時間前通電':          ('Q60','T60'),
    '冷媒漏洩点検':          ('Q61','T61'),
    'インターロック動作確認':('Q62','T62'),
}

def safe_set(ws, coord, value):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for mc in ws.merged_cells.ranges:
            if coord in mc:
                ws.cell(mc.min_row, mc.min_col).value = value
                return
    else:
        cell.value = value

def write_sheet(ws, data, unit_start, unit_end):
    """1シート分のデータを書き込む"""
    # 基本情報
    if data.get('現場名'):     safe_set(ws, 'D2', data['現場名'])
    if data.get('住所'):       safe_set(ws, 'D3', data['住所'])
    if data.get('作業日時'):
        dt = data['作業日時'].replace('T', ' ')  # 2025-06-01T09:00 → 2025-06-01 09:00
        safe_set(ws, 'T3', dt)
    if data.get('作業者'):     safe_set(ws, 'T4', data['作業者'])
    if data.get('系統名'):     safe_set(ws, 'E5', data['系統名'])
    if data.get('型式'):       safe_set(ws, 'F6', data['型式'])
    if data.get('運転状態'):   safe_set(ws, 'E7', data['運転状態'])
    if data.get('設定温度'):   safe_set(ws, 'F7', str(data['設定温度']) + '℃')
    if data.get('連結台数'):   safe_set(ws, 'T7', data['連結台数'])
    if data.get('設置年月日'): safe_set(ws, 'E8', data['設置年月日'])
    if data.get('冷媒種類'):   safe_set(ws, 'I8', data['冷媒種類'])
    if data.get('作成者'):     safe_set(ws, 'T70', data['作成者'])

    for k, cell in {'試運転':'D4','定期点検':'F4','簡易点検':'H4','修理':'J4','整備':'L4','故障判定':'N4'}.items():
        safe_set(ws, cell, ('■' if data.get('作業区分') == k else '□') + k)
    for k, cell in {'定期点検':'Q8','簡易点検':'T8'}.items():
        safe_set(ws, cell, ('■' if data.get('フロン区分') == k else '□') + k)

    # 熱源機データ（シート内のローカル列インデックス 0〜3）
    units = data.get('熱源機', [])
    for local_idx, unit_idx in enumerate(range(unit_start, unit_end)):
        if unit_idx >= len(units): break
        u = units[unit_idx]
        cols = UNIT_COLS[local_idx]
        info_col = cols['info']
        no_col   = cols['no_val']

        if u.get('No'):         safe_set(ws, f'{no_col}11',   u['No'])
        if u.get('UC製造番号'): safe_set(ws, f'{info_col}12', u['UC製造番号'])
        if u.get('冷却加熱'):   safe_set(ws, f'{info_col}13', u['冷却加熱'])
        if u.get('製造年'):     safe_set(ws, f'{info_col}14', u['製造年'])
        if u.get('記録時間'):   safe_set(ws, f'{info_col}49', u['記録時間'])

        for field, row, is_circuit in FIELD_ROWS:
            if is_circuit:
                for c in ['A','B','C','D']:
                    val = u.get(f'{field}_{c}')
                    if val is not None: safe_set(ws, f'{cols[c]}{row}', val)
            else:
                val = u.get(f'{field}_A') or u.get(field)
                if val is not None: safe_set(ws, f'{cols["A"]}{row}', val)

    # チェック結果
    checks = data.get('チェック結果', {})
    for group, items in CHECK_CELLS.items():
        selected = checks.get(group, [])
        for item, cell in items.items():
            safe_set(ws, cell, ('■' if item in selected else '□') + item)

    # 事前点検
    prechecks = data.get('試運転事前点検', {})
    for item, (ok_cell, ng_cell) in PRECHECK_MAP.items():
        result = prechecks.get(item, '')
        safe_set(ws, ok_cell, '■ＯＫ' if result == 'OK' else '□ＯＫ')
        safe_set(ws, ng_cell, '■NG'   if result == 'NG' else '□NG')

    if data.get('備考'): safe_set(ws, 'B71', data['備考'])

def generate_excel(data: dict) -> str:
    template = os.path.join(BASE, 'template.xlsx')
    out_dir  = os.path.join(os.path.expanduser('~'), 'Desktop', 'USX運転記録')
    os.makedirs(out_dir, exist_ok=True)
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    site = data.get('現場名', '').replace('/', '_').replace(' ', '_')[:20]
    out_path = os.path.join(out_dir, f'運転記録表_{site}_{ts}.xlsx')

    shutil.copy2(template, out_path)
    wb = load_workbook(out_path)

    # シートが足りなければコピーして追加
    sheet_ranges = [(0,4),(4,8),(8,12)]
    while len(wb.worksheets) < len(sheet_ranges):
        src = wb.worksheets[0]
        wb.copy_worksheet(src)

    for sheet_idx, (start, end) in enumerate(sheet_ranges):
        if sheet_idx >= len(wb.worksheets): break
        write_sheet(wb.worksheets[sheet_idx], data, start, end)

    wb.save(out_path)
    return out_path

# ===== API =====
@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/api/records', methods=['GET'])
def get_records():
    try:
        conn = get_db()
        rows = conn.execute('SELECT id, name, created_at FROM records ORDER BY created_at DESC').fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/records/<record_id>', methods=['GET'])
def get_record(record_id):
    try:
        conn = get_db()
        row = conn.execute('SELECT id, name, created_at, data FROM records WHERE id=?', (record_id,)).fetchone()
        conn.close()
        if not row: return jsonify({'error': 'Not found'}), 404
        r = dict(row); r['data'] = json.loads(r['data'])
        return jsonify(r)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/records', methods=['POST'])
def save_record():
    try:
        body = request.get_json(force=True)
        name = body.get('name', '').strip()
        if not name: return jsonify({'error': 'Name is required'}), 400
        rid = str(uuid.uuid4())
        now = datetime.now().isoformat()
        conn = get_db()
        conn.execute('INSERT INTO records (id,name,created_at,data) VALUES (?,?,?,?)',
                     (rid, name, now, json.dumps(body.get('data', {}))))
        conn.commit(); conn.close()
        return jsonify({'id': rid, 'name': name, 'created_at': now}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/records/<record_id>', methods=['DELETE'])
def delete_record(record_id):
    try:
        conn = get_db()
        conn.execute('DELETE FROM records WHERE id=?', (record_id,))
        conn.commit(); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate', methods=['POST'])
def generate():
    try:
        data = request.get_json(force=True)
        out_path = generate_excel(data)
        return send_file(out_path, as_attachment=True,
                         download_name=os.path.basename(out_path),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== 起動 =====
def open_browser(port):
    import time; time.sleep(1.2)
    webbrowser.open(f'http://localhost:{port}')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5199))
    is_render = 'RENDER' in os.environ
    print('='*50)
    print(' USX 運転記録表サーバー 起動中...')
    print('='*50)
    if not is_render:
        import socket
        try: ip = socket.gethostbyname(socket.gethostname())
        except: ip = '127.0.0.1'
        print(f' PC用URL:     http://localhost:{port}')
        print(f' スマホ用URL: http://{ip}:{port}')
        threading.Thread(target=open_browser, args=(port,), daemon=True).start()
    app.run(host='0.0.0.0', port=port, debug=False)

# ===== 写真API =====

def resize_image(img_bytes, max_w=800, max_h=600):
    """画像をリサイズしてJPEGバイト列で返す"""
    img = PILImage.open(io.BytesIO(img_bytes))
    img = img.convert('RGB')
    img.thumbnail((max_w, max_h), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=85)
    return buf.getvalue()

@app.route('/api/photos/<record_id>', methods=['GET'])
def get_photos(record_id):
    try:
        conn = get_db()
        rows = conn.execute(
            'SELECT id, filename, caption, created_at FROM photos WHERE record_id=? ORDER BY created_at',
            (record_id,)).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/photos/<record_id>', methods=['POST'])
def upload_photo(record_id):
    try:
        body = request.get_json(force=True)
        img_b64  = body.get('image', '')   # base64 data URL
        caption  = body.get('caption', '')
        if not img_b64:
            return jsonify({'error': 'No image'}), 400

        # base64デコード
        if ',' in img_b64:
            img_b64 = img_b64.split(',', 1)[1]
        img_bytes = base64.b64decode(img_b64)
        img_bytes = resize_image(img_bytes)

        pid = str(uuid.uuid4())
        filename = f'{pid}.jpg'
        filepath = os.path.join(IMG_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(img_bytes)

        now = datetime.now().isoformat()
        conn = get_db()
        conn.execute(
            'INSERT INTO photos (id, record_id, filename, caption, created_at) VALUES (?,?,?,?,?)',
            (pid, record_id, filename, caption, now))
        conn.commit(); conn.close()
        return jsonify({'id': pid, 'caption': caption, 'created_at': now}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/photos/item/<photo_id>', methods=['DELETE'])
def delete_photo(photo_id):
    try:
        conn = get_db()
        row = conn.execute('SELECT filename FROM photos WHERE id=?', (photo_id,)).fetchone()
        if row:
            filepath = os.path.join(IMG_DIR, row['filename'])
            if os.path.exists(filepath): os.remove(filepath)
            conn.execute('DELETE FROM photos WHERE id=?', (photo_id,))
            conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/photos/file/<photo_id>', methods=['GET'])
def get_photo_file(photo_id):
    try:
        conn = get_db()
        row = conn.execute('SELECT filename FROM photos WHERE id=?', (photo_id,)).fetchone()
        conn.close()
        if not row: return jsonify({'error': 'Not found'}), 404
        return send_file(os.path.join(IMG_DIR, row['filename']), mimetype='image/jpeg')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_with_photos', methods=['POST'])
def generate_with_photos():
    try:
        body = request.get_json(force=True)
        data = body.get('data', {})
        record_id = body.get('record_id', '')

        # 運転記録Excel生成
        template = os.path.join(BASE, 'template.xlsx')
        photo_template = os.path.join(BASE, 'photo_template.xlsx')
        out_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'USX運転記録')
        os.makedirs(out_dir, exist_ok=True)
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        site = data.get('現場名', '').replace('/', '_').replace(' ', '_')[:20]
        out_path = os.path.join(out_dir, f'運転記録表_{site}_{ts}.xlsx')

        shutil.copy2(template, out_path)
        wb = load_workbook(out_path)

        sheet_ranges = [(0,4),(4,8),(8,12)]
        while len(wb.worksheets) < len(sheet_ranges):
            wb.copy_worksheet(wb.worksheets[0])

        for sheet_idx, (start, end) in enumerate(sheet_ranges):
            if sheet_idx >= len(wb.worksheets): break
            write_sheet(wb.worksheets[sheet_idx], data, start, end)

        # 写真台帳を追加
        if record_id and os.path.exists(photo_template):
            conn = get_db()
            photos = conn.execute(
                'SELECT id, filename, caption, created_at FROM photos WHERE record_id=? ORDER BY created_at',
                (record_id,)).fetchall()
            conn.close()

            if photos:
                wb_photo = load_workbook(photo_template)
                photos_per_sheet = 3
                sheet_idx = 0

                for i, photo in enumerate(photos):
                    local_idx = i % photos_per_sheet
                    if local_idx == 0 and i > 0:
                        sheet_idx += 1

                    if sheet_idx >= len(wb_photo.worksheets) - 1:
                        break

                    ws_p = wb_photo.worksheets[sheet_idx]
                    filepath = os.path.join(IMG_DIR, photo['filename'])
                    if not os.path.exists(filepath): continue

                    # 写真エリアの行: 1枚目=3-23行, 2枚目=25-45行, 3枚目=47-67行
                    photo_rows = [(3, 23), (25, 45), (47, 67)]
                    info_rows  = [
                        {'no': 3, 'date': 4, 'author': 5, 'place': 6, 'content_start': 7},
                        {'no': 25, 'date': 26, 'author': 27, 'place': 28, 'content_start': 29},
                        {'no': 47, 'date': 48, 'author': 49, 'place': 50, 'content_start': 51},
                    ]

                    pr = photo_rows[local_idx]
                    ir = info_rows[local_idx]

                    # 写真を貼り付け
                    try:
                        xl_img = XLImage(filepath)
                        # B列の幅から画像サイズを計算（約150px幅）
                        xl_img.width  = 380
                        xl_img.height = 285
                        col_letter = 'B'
                        anchor = f'{col_letter}{pr[0]}'
                        ws_p.add_image(xl_img, anchor)
                    except Exception:
                        pass

                    # キャプション・撮影日を書き込み
                    created = photo['created_at'][:10] if photo['created_at'] else ''
                    caption = photo['caption'] or ''
                    ws_p.cell(ir['date'], 8, created)
                    ws_p.cell(ir['content_start'], 8, caption)

                # 写真シートをメインワークブックにコピー
                for ws_p in wb_photo.worksheets:
                    if ws_p.title == '写真 (追加用)': continue
                    ws_new = wb.create_sheet(title=ws_p.title)
                    for row in ws_p.iter_rows():
                        for cell in row:
                            new_cell = ws_new.cell(cell.row, cell.column, cell.value)
                            if cell.has_style:
                                new_cell.font      = cell.font.copy()
                                new_cell.border    = cell.border.copy()
                                new_cell.fill      = cell.fill.copy()
                                new_cell.alignment = cell.alignment.copy()
                    for mc in ws_p.merged_cells.ranges:
                        ws_new.merge_cells(str(mc))
                    for img in ws_p._images:
                        ws_new.add_image(img)
                    for col in ws_p.column_dimensions:
                        ws_new.column_dimensions[col].width = ws_p.column_dimensions[col].width
                    for row in ws_p.row_dimensions:
                        ws_new.row_dimensions[row].height = ws_p.row_dimensions[row].height

        wb.save(out_path)
        return send_file(out_path, as_attachment=True,
                         download_name=os.path.basename(out_path),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
